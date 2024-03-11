import paramiko
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('./file/setting.xlsx')
ws = wb.active # wb['工作表1']
for row in range(1, ws.max_row):
    d = []
    [d.append(col[row].value) for col in ws.iter_cols(1, ws.max_column)]

    username = d[0]
    password = d[1]
    host = d[2]
    port = d[3]
    command = d[4]
    executed = d[5]

    # get ssh setting info
    # r_index = str(row + 1)
    # username = ws[f'{get_column_letter(1) + r_index}'].value
    # password = ws[f'{get_column_letter(2) + r_index}'].value
    # host = ws[f'{get_column_letter(3) + r_index}'].value
    # port = ws[f'{get_column_letter(4) + r_index}'].value
    # command = ws[f'{get_column_letter(5) + r_index}'].value

    if str(executed).lower() not in ('yes', 'y'): continue

    try:
        # connect
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(host, port, username, password)

        # command = '''
        # cd /var/www/html
        # ls
        # '''


        '''
        # --- all command in one process ---
        stdin, stdout, stderr = ssh.exec_command(command)

        # read
        result = stdout.read().decode('utf-8')

        # show
        print(f'Host:{host} command was executed successfully, execution results as follows:\n{result}')
        # --- ---
        '''

        # --- each command processing ---
        stdin, stdout, stderr = ssh.exec_command(command, get_pty=True)

        for line in iter(stdout.readline, ''):
            print(line, end='')

        print(f'Host:{host} command was executed successfully', end='\n\n')
        # --- ---

    except paramiko.AuthenticationException:
        print('Authentication failed. Please check your username and password.')

    except paramiko.SSHException as e:
        print(f'Unable to establish SSH connection: {e}')

    except Exception as e:
        print(f'An unexpected error occurred: {e}')

    finally:
        # disconnect
        if ssh:
            ssh.close()
